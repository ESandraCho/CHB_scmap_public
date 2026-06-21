#!/usr/bin/env python3
"""build_supplementary_tables.py — assemble the manuscript Supplementary Tables 1-3
into a single formatted Excel workbook (one sheet per table).

This is the Excel-export layer over the CSVs that `supp_dataset_analysis_table.py` emits,
so the two stay consistent (single source of truth):
  - Supplementary Table 1 (datasets):  results/tables/supp_datasets_provenance.csv
      (the 7-dataset provenance table; this script AUGMENTS it with two computed cell-count
       columns — object total and cells used — read live from the data, not hard-coded)
  - Supplementary Table 2 (analysis -> dataset/figure map): results/tables/supp_analysis_dataset_map.csv
      (read verbatim)
  - Supplementary Table 3 (channel classification): results/tables/channelopathy_classification.tsv

Run `supp_dataset_analysis_table.py` first (it writes the two supp_*.csv inputs).

Output: results/tables/CHB_supplementary_tables.xlsx  (Table S1 / S2 / S3 sheets).
No data values are hard-coded; every value is read from a committed CSV/TSV or computed from
the raw data.
"""
import csv
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
TABLES = HERE.parent / "results" / "tables"
RAW = HERE.parent.parent / "private" / "data"
OUT = TABLES / "CHB_supplementary_tables.xlsx"

# ----------------------------------------------------------------------------- styling
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=12, color="1F3864")
CELL_FONT = Font(size=9)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")


def _read_csv(path, skip_comment=True):
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        for line in f:
            if skip_comment and line.startswith("#"):
                continue
            rows.append(line.rstrip("\n"))
    return list(csv.reader(rows))


def _read_tsv(path):
    with open(path, newline="", encoding="utf-8") as f:
        return [r for r in csv.reader(f, delimiter="\t") if any(c.strip() for c in r)]


def _write_sheet(wb, sheet_name, title, header, data_rows, widths):
    ws = wb.create_sheet(sheet_name)
    ws.cell(1, 1, title).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
    hrow = 3
    for c, h in enumerate(header, 1):
        cell = ws.cell(hrow, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = BORDER
    for r, row in enumerate(data_rows, hrow + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.font = CELL_FONT
            cell.alignment = WRAP
            cell.border = BORDER
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(hrow + 1, 1)
    return ws


# ----------------------------------------------------------- cell-count computation (S1)
def _gz_header_ncols(path, sep):
    """Cells = data columns in a gzipped genes x cells matrix (first column = gene index)."""
    import gzip
    if not path.exists():
        return None
    with gzip.open(path, "rt") as f:
        return len(f.readline().rstrip("\n").split(sep)) - 1


def _gz_nrows(path):
    """Cells = data rows (minus header) in a gzipped per-cell info table."""
    import gzip
    if not path.exists():
        return None
    with gzip.open(path, "rt") as f:
        return sum(1 for _ in f) - 1


def _lazar_object_total(lazar_dir):
    """Total cells across all Lazar HDCA shoji lanes, from each .h5 Expression shape
    (cells x genes; shape only). None if h5py or files unavailable."""
    import glob
    files = sorted(glob.glob(str(lazar_dir / "*.h5")))
    if not files:
        return None
    try:
        import h5py
    except ImportError:
        return None
    return sum(int(h5py.File(fp, "r")["shoji"]["Expression"].shape[0]) for fp in files)


def _cell_counts_by_dataset():
    """Map dataset-name -> (cells_in_object, cells_used), keyed to the dataset names used
    in supp_datasets_provenance.csv. Counts are read live (committed outputs + raw matrices),
    never hard-coded; missing pieces fall back to "" so the build never fails on absent data.
    """
    counts = {}

    # h5ad-backed datasets (Lim SAN/AVN, Kanemaru, Sim): object + used counts already
    # computed live by node_datasets_methods_table.py
    nd = _read_csv(TABLES / "node_datasets_methods_summary.csv")
    if nd:
        cols = {h: i for i, h in enumerate(nd[0])}
        for r in nd[1:]:
            counts[r[cols["dataset"]]] = (r[cols["n_cells_object"]], r[cols["n_cells_used"]])

    # Cui / Wang (raw STRT matrices): used counts from the vulnerability analysis output;
    # object totals from the raw matrix dimensions
    vuln = json.loads((TABLES / "fetal_vs_adult_vulnerability.json").read_text())
    nc = vuln["n_cells"]
    cui_obj = _gz_header_ncols(
        RAW / "cui2019_gse106118" / "GSE106118_UMI_count_merge.txt.gz", "\t")
    wang_obj = _gz_nrows(
        RAW / "wang2020_gse109816" / "GSE109816_normal_heart_cell_info.txt.gz")
    counts["Cui 2019 (fetal STRT)"] = (
        str(cui_obj) if cui_obj is not None else "",
        str(int(nc["cui_early_fetal"]) + int(nc["cui_late_fetal"])))
    counts["Wang 2020 (adult STRT)"] = (
        str(wang_obj) if wang_obj is not None else "", str(int(nc["wang_adult"])))

    # Lazar (HDCA): object total summed across lanes; used = CM passing the per-donor gate
    lazar = [r for r in _read_csv(TABLES / "lazar_dev_replication.csv")
             if r and r[0] != "donor"]
    lazar_obj = _lazar_object_total(RAW / "lazar2025")
    lazar_used = sum(int(r[2]) for r in lazar) if lazar else ""
    counts["Lázár 2025 (HDCA first-trimester)"] = (
        str(lazar_obj) if lazar_obj is not None else "", str(lazar_used))
    return counts


def build():
    prov = TABLES / "supp_datasets_provenance.csv"
    amap = TABLES / "supp_analysis_dataset_map.csv"
    if not prov.exists() or not amap.exists():
        sys.exit(f"missing input(s): run supp_dataset_analysis_table.py first "
                 f"({prov.name} / {amap.name})")

    wb = Workbook()
    wb.remove(wb.active)

    # ---- Supplementary Table 1: datasets (provenance CSV + computed counts) ----
    ds = _read_csv(prov)
    ds_header_raw, ds_prov_rows = ds[0], ds[1:]
    counts = _cell_counts_by_dataset()
    ds_rows = [row + list(counts.get(row[0], ("", ""))) for row in ds_prov_rows]
    # header = provenance columns (renamed) + the two appended count columns
    ds_header = ["Dataset", "Object / directory", "Accession", "Tissue / stage", "Chemistry",
                 "Platform used (this study)", "Donors", "Citation",
                 "Cells in object", "Cells used"]
    if len(ds_header) != len(ds_header_raw) + 2:
        sys.exit(f"provenance CSV has {len(ds_header_raw)} cols; expected 8 (+2 counts)")
    _write_sheet(
        wb, "Table S1 - datasets",
        "Supplementary Table 1. Single-cell / single-nucleus datasets included in this study.",
        ds_header, ds_rows, widths=[26, 30, 34, 30, 22, 40, 22, 30, 14, 12])

    # ---- Supplementary Table 2: analysis -> dataset/figure map (CSV verbatim) ----
    am = _read_csv(amap)
    am_header_raw, am_rows = am[0], am[1:]
    am_header = ["Analysis", "Datasets used", "Cell types compared",
                 "Method", "Comparison type", "Figure(s)"]
    if len(am_header) != len(am_header_raw):
        sys.exit(f"analysis-map CSV has {len(am_header_raw)} cols; expected 6")
    _write_sheet(
        wb, "Table S2 - analysis map",
        "Supplementary Table 2. Analysis-to-dataset map: which analysis used which datasets, "
        "cell types, method, comparison type and figures.",
        am_header, am_rows, widths=[34, 38, 40, 40, 30, 22])

    # ---- Supplementary Table 3: channel classification (TSV) ----
    cls = _read_tsv(TABLES / "channelopathy_classification.tsv")
    cls_header_raw, cls_rows = cls[0], cls[1:]
    cls_header = ["Gene", "Channel / current", "Antibody", "Autoimmune class",
                  "Arrhythmia class", "Life stage", "ECG symptom", "EP mechanism",
                  "Inherited counterpart", "References"]
    _write_sheet(
        wb, "Table S3 - classification",
        "Supplementary Table 3. Autoimmune-arrhythmia channel classification "
        "(reported clinical phenotype of each channel autoantibody). The Figure-1 encoding "
        "is derived entirely from this table.",
        cls_header, cls_rows, widths=[10, 24, 24, 30, 14, 16, 26, 44, 28, 44])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  Table S1: {len(ds_rows)} datasets")
    print(f"  Table S2: {len(am_rows)} analyses")
    print(f"  Table S3: {len(cls_rows)} classification rows")


if __name__ == "__main__":
    build()
